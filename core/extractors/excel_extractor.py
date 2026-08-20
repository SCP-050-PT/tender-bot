"""
Экстрактор текста из Excel файлов (XLSX / XLS).
Багфиксы v6.6-r2:
  - Корректное извлечение количества из «Обоснования НМЦК»
  - Поиск количества даже без колонки «услуга»
  - Фильтрация фантомных чисел (телефоны, адреса)

v7.1.0: Извлечение цены за единицу из обоснований НМЦК (openpyxl + xlrd)
v7.2.0: Пропуск merged cells, нормализация headers, фильтр ложного qty,
         ограничение поиска заголовков первыми 15 строками, убран spam-лог
v7.2.1: sheet.nrows вместо sheet.max_row (xlrd), debug-лог только при price_col=-1
"""

import re
from pathlib import Path
from typing import List, Tuple, Optional
from loguru import logger

from core.config.document_config import (
    QUANTITY_COLUMN_KEYWORDS,
    SERVICE_ROW_KEYWORDS,
    NMCK_FILE_KEYWORDS,
    UNIT_PRICE_COLUMN_KEYWORDS,
)
from core.extractors.base_extractor import BaseExtractor

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd

    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False


class ExcelExtractor(BaseExtractor):
    """Извлекает текст из Excel файлов со структурированными данными."""

    SUPPORTED_EXTENSIONS = ["xlsx", "xls"]

    # Паттерны фантомных чисел (телефоны, ИНН, КПП, адреса)
    PHANTOM_PATTERNS = [
        re.compile(r"^\+?\d[\d\s\-()]{7,15}$"),  # телефон
        re.compile(r"^\d{10,12}$"),  # ИНН/КПП/ОГРН
        re.compile(r"^\d{6}$"),  # почтовый индекс
        re.compile(r"^\d{1,3}[-/]\d{1,3}$"),  # дроби типа 12/34
    ]

    def extract(self, file_path: Path, doc_name: str = "") -> str:
        is_xlsx, is_xls = self._detect_format(file_path)

        if HAS_OPENPYXL and is_xlsx:
            return self._extract_structured_openpyxl(file_path, doc_name)
        elif HAS_XLRD and is_xls:
            return self._extract_structured_xlrd(file_path, doc_name)
        else:
            logger.warning(f"[ExcelExtractor] Нет библиотеки для {file_path.suffix}")
            return self._extract_fallback(file_path)

    def _detect_format(self, file_path: Path) -> Tuple[bool, bool]:
        """Определяет формат Excel по магическим байтам и расширению."""
        is_xlsx = False
        is_xls = False

        with open(file_path, "rb") as f:
            header = f.read(8)
            if header.startswith(b"\x50\x4b\x03\x04"):
                is_xlsx = True
            elif header.startswith(b"\xd0\xcf\x11\xe0"):
                is_xls = True

        # Fallback на расширение
        suffix = file_path.suffix.lower()
        if not is_xlsx and not is_xls:
            if suffix == ".xlsx":
                is_xlsx = True
            elif suffix == ".xls":
                is_xls = True

        return is_xlsx, is_xls

    # ================================================================
    # OPENPYXL (XLSX)
    # ================================================================

    def _extract_structured_openpyxl(self, file_path: Path, doc_name: str) -> str:
        """Извлекает XLSX со структурированными данными."""
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            all_texts = []
            extracted_quantities = []

            for sheet in wb.worksheets:
                sheet_texts, quantities = self._process_sheet_openpyxl(sheet, doc_name)
                all_texts.extend(sheet_texts)
                extracted_quantities.extend(quantities)

            if extracted_quantities:
                prefix = "\n".join(
                    f"=== ИЗВЛЕЧЕНО ИЗ ТАБЛИЦЫ: Количество = {q} ==="
                    for q in extracted_quantities[:5]
                )
                all_texts.insert(0, prefix)

            result = "\n".join(all_texts)
            logger.info(
                f"[ExcelExtractor] XLSX извлечён: {len(all_texts)} строк, "
                f"количеств найдено: {len(extracted_quantities)}"
            )
            return result

        except Exception as e:
            logger.error(f"[ExcelExtractor] Ошибка openpyxl: {e}")
            return self._extract_fallback(file_path)

    def _process_sheet_openpyxl(
        self, sheet, doc_name: str
    ) -> Tuple[List[str], List[int]]:
        """Обрабатывает один лист XLSX. Возвращает (тексты, количества).
        v7.1.0: Также извлекает цену за единицу из обоснований НМЦК.
        v7.2.0: Пропускает merged cells, нормализация headers,
                 ограничение поиска заголовков первыми 15 строками.
        """
        sheet_texts = []
        extracted_quantities = []
        headers = []
        quantity_col_idx = -1
        service_col_idx = -1
        unit_price_col_idx = -1
        is_nmck_file = self._is_nmck_file(doc_name)

        for row_idx, row in enumerate(sheet.iter_rows()):
            row_values = [
                str(cell.value) if cell.value is not None else "" for cell in row
            ]
            row_text = [v for v in row_values if v.strip()]

            if not row_text:
                continue

            # --- Поиск заголовков в первых 15 строках ---
            if row_idx < 15 and not headers:
                unique_values = set(v.lower().strip() for v in row_values if v.strip())

                # Пропускаем merged cells (1 уникальное значение)
                if len(unique_values) <= 1:
                    continue

                # Проверяем ключевые слова заголовков
                row_lower = " ".join(v.lower() for v in row_values)
                has_header_keywords = any(
                    kw in row_lower
                    for kw in [
                        "наименование",
                        "кол-во",
                        "количество",
                        "цена",
                        "окпд",
                        "единица",
                        "№",
                        "номер",
                    ]
                )

                if not has_header_keywords:
                    continue

                # Нормализация headers (убираем \n, лишние пробелы)
                headers = [re.sub(r"\s+", " ", v.lower()).strip() for v in row_values]
                quantity_col_idx = self._find_quantity_column(headers)
                service_col_idx = self._find_service_column(headers)
                unit_price_col_idx = self._find_unit_price_column(headers)
                logger.info(
                    f"[ExcelExtractor] Заголовки на строке {row_idx}: "
                    f"qty={quantity_col_idx}, svc={service_col_idx}, "
                    f"price={unit_price_col_idx}"
                )
                # Debug только если цена не найдена
                if unit_price_col_idx == -1:
                    logger.debug(
                        f"[ExcelExtractor] price_col=-1. Заголовки: {headers[:20]}"
                    )
                continue

            # После 15 строк без заголовков — прекращаем поиск
            if row_idx >= 14 and not headers:
                logger.debug(
                    f"[ExcelExtractor] Заголовки не найдены в первых 15 строках, "
                    f'пропускаем оставшиеся строки листа "{sheet.title}"'
                )
                break

            # --- Обработка строк данных (заголовки уже найдены) ---

            # Формируем строку с подписями заголовков
            enriched_row = self._enrich_row(row_values, headers, quantity_col_idx)
            if enriched_row:
                sheet_texts.append(" | ".join(enriched_row))

            # Ищем количество
            qty = self._extract_quantity_from_row(
                row_values, headers, quantity_col_idx, service_col_idx, is_nmck_file
            )
            if qty is not None and qty not in extracted_quantities:
                extracted_quantities.append(qty)

            # Извлекаем цену за единицу из обоснования НМЦК
            if is_nmck_file and unit_price_col_idx >= 0:
                price = self._extract_unit_price_from_row(
                    row_values, unit_price_col_idx, service_col_idx
                )
                if price is not None:
                    prefix_line = (
                        f"=== ЦЕНА ЗА ЕДИНИЦУ ИЗ ОБОСНОВАНИЯ НМЦК: {price:.2f} ₽ ==="
                    )
                    if prefix_line not in sheet_texts:
                        sheet_texts.insert(0, prefix_line)
                        logger.info(
                            f"[ExcelExtractor v7.1.0] Цена за ед. из НМЦК: {price:.2f} ₽"
                        )

        if sheet_texts:
            sheet_texts.insert(0, f"=== ЛИСТ: {sheet.title} ===")

        return sheet_texts, extracted_quantities

    # ================================================================
    # XLRD (XLS)
    # ================================================================

    def _extract_structured_xlrd(self, file_path: Path, doc_name: str) -> str:
        """Извлекает XLS со структурированными данными."""
        try:
            wb = xlrd.open_workbook(file_path)
            all_texts = []
            extracted_quantities = []

            for sheet in wb.sheets():
                sheet_texts, quantities = self._process_sheet_xlrd(sheet, doc_name)
                all_texts.extend(sheet_texts)
                extracted_quantities.extend(quantities)

            if extracted_quantities:
                prefix = "\n".join(
                    f"=== ИЗВЛЕЧЕНО ИЗ ТАБЛИЦЫ: Количество = {q} ==="
                    for q in extracted_quantities[:5]
                )
                all_texts.insert(0, prefix)

            result = "\n".join(all_texts)
            logger.info(
                f"[ExcelExtractor] XLS извлечён: {len(all_texts)} строк, "
                f"количеств найдено: {len(extracted_quantities)}"
            )
            return result

        except Exception as e:
            logger.error(f"[ExcelExtractor] Ошибка xlrd: {e}")
            return self._extract_fallback(file_path)

    def _process_sheet_xlrd(self, sheet, doc_name: str) -> Tuple[List[str], List[int]]:
        """Обрабатывает один лист XLS.
        v7.1.0: Также извлекает цену за единицу из обоснований НМЦК.
        v7.2.0: Пропускает merged cells, ограничение 15 строк.
        v7.2.1: sheet.nrows вместо sheet.max_row (xlrd не имеет max_row).
        """
        sheet_texts = []
        extracted_quantities = []
        headers = []
        quantity_col_idx = -1
        service_col_idx = -1
        unit_price_col_idx = -1
        is_nmck_file = self._is_nmck_file(doc_name)

        for row_idx in range(sheet.nrows):
            row_values = [
                str(sheet.cell_value(row_idx, col_idx))
                for col_idx in range(sheet.ncols)
            ]
            row_text = [v for v in row_values if v.strip()]

            if not row_text:
                continue

            # --- Поиск заголовков в первых 15 строках ---
            if row_idx < 15 and not headers:
                unique_values = set(v.lower().strip() for v in row_values if v.strip())

                # Пропускаем merged cells (1 уникальное значение)
                if len(unique_values) <= 1:
                    continue

                # Проверяем ключевые слова заголовков
                row_lower = " ".join(v.lower() for v in row_values)
                has_header_keywords = any(
                    kw in row_lower
                    for kw in [
                        "наименование",
                        "кол-во",
                        "количество",
                        "цена",
                        "окпд",
                        "единица",
                        "№",
                        "номер",
                    ]
                )

                if not has_header_keywords:
                    continue

                # Нормализация headers
                headers = [re.sub(r"\s+", " ", v.lower()).strip() for v in row_values]
                quantity_col_idx = self._find_quantity_column(headers)
                service_col_idx = self._find_service_column(headers)
                unit_price_col_idx = self._find_unit_price_column(headers)
                logger.info(
                    f"[ExcelExtractor] Заголовки на строке {row_idx}: "
                    f"qty={quantity_col_idx}, svc={service_col_idx}, "
                    f"price={unit_price_col_idx}"
                )
                # Debug только если цена не найдена
                if unit_price_col_idx == -1:
                    logger.debug(
                        f"[ExcelExtractor] price_col=-1. Заголовки: {headers[:20]}"
                    )
                continue

            # После 15 строк без заголовков — прекращаем поиск
            # v7.2.1: sheet.nrows вместо sheet.max_row (xlrd не имеет max_row)
            if row_idx >= 14 and not headers:
                logger.debug(
                    f"[ExcelExtractor] Заголовки не найдены в первых 15 строках, "
                    f'пропускаем оставшиеся {sheet.nrows - 15} строк листа "{sheet.name}"'
                )
                break

            # --- Обработка строк данных ---

            enriched_row = self._enrich_row(row_values, headers, quantity_col_idx)
            if enriched_row:
                sheet_texts.append(" | ".join(enriched_row))

            qty = self._extract_quantity_from_row(
                row_values, headers, quantity_col_idx, service_col_idx, is_nmck_file
            )
            if qty is not None and qty not in extracted_quantities:
                extracted_quantities.append(qty)

            if is_nmck_file and unit_price_col_idx >= 0:
                price = self._extract_unit_price_from_row(
                    row_values, unit_price_col_idx, service_col_idx
                )
                if price is not None:
                    prefix_line = (
                        f"=== ЦЕНА ЗА ЕДИНИЦУ ИЗ ОБОСНОВАНИЯ НМЦК: {price:.2f} ₽ ==="
                    )
                    if prefix_line not in sheet_texts:
                        sheet_texts.insert(0, prefix_line)
                        logger.info(
                            f"[ExcelExtractor v7.1.0] XLS цена за ед. из НМЦК: {price:.2f} ₽"
                        )

        if sheet_texts:
            sheet_texts.insert(0, f"=== ЛИСТ: {sheet.name} ===")

        return sheet_texts, extracted_quantities

    # ================================================================
    # ВСПОМОГАТЕЛЬНЫЕ МЕТОДЫ
    # ================================================================

    def _find_quantity_column(self, headers: List[str]) -> int:
        """Находит индекс колонки с количеством."""
        for idx, h in enumerate(headers):
            if any(kw in h for kw in QUANTITY_COLUMN_KEYWORDS):
                logger.info(f"[ExcelExtractor] Колонка количества: '{h}' (idx {idx})")
                return idx
        return -1

    def _find_service_column(self, headers: List[str]) -> int:
        """Находит индекс колонки с наименованием услуги."""
        for idx, h in enumerate(headers):
            if any(kw in h for kw in ["наименование", "услуга", "работа", "предмет"]):
                return idx
        return -1

    def _find_unit_price_column(self, headers: List[str]) -> int:
        """Находит индекс колонки с ценой за единицу (v7.1.0).
        v7.2.1: Не логирует здесь — лог в вызывающем коде только при price_col=-1.
        """
        for idx, h in enumerate(headers):
            if any(kw in h for kw in UNIT_PRICE_COLUMN_KEYWORDS):
                logger.info(f"[ExcelExtractor] Колонка цены за ед.: '{h}' (idx {idx})")
                return idx
        return -1

    def _is_nmck_file(self, doc_name: str) -> bool:
        """Проверяет, является ли файл «Обоснованием НМЦК»."""
        name_lower = doc_name.lower()
        return any(kw in name_lower for kw in NMCK_FILE_KEYWORDS)

    def _enrich_row(
        self, row_values: List[str], headers: List[str], quantity_col_idx: int
    ) -> List[str]:
        """Формирует строку с подписями заголовков."""
        enriched = []
        for col_idx, cell_value in enumerate(row_values):
            if not cell_value.strip():
                continue
            if col_idx < len(headers) and headers[col_idx]:
                header = headers[col_idx]
                if col_idx == quantity_col_idx:
                    enriched.append(f"{header}: {cell_value}")
                else:
                    enriched.append(cell_value)
            else:
                enriched.append(cell_value)
        return enriched

    def _extract_quantity_from_row(
        self,
        row_values: List[str],
        headers: List[str],
        quantity_col_idx: int,
        service_col_idx: int,
        is_nmck_file: bool,
    ) -> Optional[int]:
        """
        Извлекает количество из строки.
        Багфикс v6.6-r2: работает и без колонки услуги (для «Обоснования НМЦК»).
        v7.2.0: Фильтр ложного qty (цена попала в колонку кол-во).
        """
        if quantity_col_idx < 0 or quantity_col_idx >= len(row_values):
            return None

        qty_cell = row_values[quantity_col_idx]
        if not qty_cell or not qty_cell.strip():
            return None

        qty_str = str(qty_cell).replace(" ", "").replace(",", ".")

        # Проверяем, что это не фантомное число
        if self._is_phantom_number(qty_str):
            return None

        try:
            qty = int(float(qty_str))
        except (ValueError, TypeError):
            return None

        # v7.2.0: Фильтр ложного qty — если > 1000 и рядом десятичные числа (цены)
        if qty > 1000 and quantity_col_idx >= 0:
            for adj_idx in range(
                max(0, quantity_col_idx - 2), min(len(row_values), quantity_col_idx + 3)
            ):
                if adj_idx == quantity_col_idx:
                    continue
                adj_val = row_values[adj_idx].replace(" ", "").replace(",", ".")
                if re.search(r"\d+\.\d{2}", adj_val):
                    logger.debug(
                        f"[ExcelExtractor] qty={qty} пропущено: похоже на цену"
                    )
                    return None

        # Фильтр: отбрасываем явно нереалистичные значения
        if qty <= 0 or qty > 10000:
            return None

        # Если есть колонка услуги — проверяем ключевые слова
        if service_col_idx >= 0 and service_col_idx < len(row_values):
            service_str = str(row_values[service_col_idx]).lower()
            if any(kw in service_str for kw in SERVICE_ROW_KEYWORDS):
                logger.info(
                    f"[ExcelExtractor] Найдено qty={qty} (услуга: {service_str[:50]})"
                )
                return qty

        # Для «Обоснования НМЦК» — не требуем колонку услуги
        if is_nmck_file:
            logger.info(f"[ExcelExtractor] Найдено qty={qty} (Обоснование НМЦК)")
            return qty

        # Если нет колонки услуги и это не НМЦК — всё равно возвращаем
        if service_col_idx < 0:
            logger.info(f"[ExcelExtractor] Найдено qty={qty} (без колонки услуги)")
            return qty

        return None

    def _extract_unit_price_from_row(
        self,
        row_values: List[str],
        unit_price_col_idx: int,
        service_col_idx: int,
    ) -> Optional[float]:
        """Извлекает цену за единицу из строки обоснования НМЦК (v7.1.0)."""
        if unit_price_col_idx < 0 or unit_price_col_idx >= len(row_values):
            return None

        price_cell = row_values[unit_price_col_idx]
        if not price_cell or not price_cell.strip():
            return None

        price_str = str(price_cell).replace(" ", "").replace(",", ".")

        # Убираем текст типа "руб." или "₽"
        price_str = re.sub(r"[^\d.]", "", price_str)

        if not price_str:
            return None

        try:
            price = float(price_str)
        except (ValueError, TypeError):
            return None

        # Sanity check: цена за единицу СОУТ/ПЛК/ОПР обычно 100-50000 ₽
        if price <= 0 or price > 100000:
            return None

        # Проверяем, что строка содержит релевантную услугу
        if service_col_idx >= 0 and service_col_idx < len(row_values):
            service_str = str(row_values[service_col_idx]).lower()
            if any(kw in service_str for kw in SERVICE_ROW_KEYWORDS):
                return price

        return price

    def _is_phantom_number(self, text: str) -> bool:
        """Проверяет, является ли строка фантомным числом (телефон, ИНН и т.д.)."""
        for pattern in self.PHANTOM_PATTERNS:
            if pattern.match(text):
                return True
        return False

    def _extract_fallback(self, file_path: Path) -> str:
        """Fallback: обычное извлечение как plain text."""
        try:
            if HAS_OPENPYXL:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                texts = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        row_text = [
                            str(cell.value) for cell in row if cell.value is not None
                        ]
                        if row_text:
                            texts.append(" | ".join(row_text))
                return "\n".join(texts)
            elif HAS_XLRD:
                wb = xlrd.open_workbook(file_path)
                texts = []
                for sheet in wb.sheets():
                    for row_idx in range(sheet.nrows):
                        row_text = [
                            str(sheet.cell_value(row_idx, col_idx))
                            for col_idx in range(sheet.ncols)
                        ]
                        if row_text:
                            texts.append(" | ".join(row_text))
                return "\n".join(texts)
            return ""
        except Exception as e:
            logger.error(f"[ExcelExtractor] Ошибка fallback: {e}")
            return ""
