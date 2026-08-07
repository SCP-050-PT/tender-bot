"""
core/documents/excel_extractor.py
Извлечение текста из Excel-файлов (XLSX/XLS).
Вынесено из document_processor.py (v6.5).
"""

import re
from pathlib import Path
from loguru import logger

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

QUANTITY_KEYWORDS = [
    "количество", "кол-во", "кол.", "объем", "объём",
    "планируемое кол-во", "кол-во обучаемых", "обучаемых",
    "чел.", "человек", "слушателей",
    "количество рабочих мест", "кол-во рабочих мест",
]

SERVICE_KEYWORDS = [
    "соут", "специальная оценка", "оценка условий труда",
    "обучение", "обучению", "повышение квалификации",
    "плк", "лабораторный контроль",
    "опр", "оценка профессиональных рисков",
]


class ExcelExtractor:
    """Извлекает текст из Excel с структурированными данными."""

    def extract(self, file_path: Path, doc_name: str) -> str:
        """Определяет тип Excel и извлекает структурированно."""
        is_xlsx, is_xls = self._detect_format(file_path)

        if HAS_OPENPYXL and is_xlsx:
            return self._extract_structured_xlsx(file_path, doc_name)
        elif HAS_XLRD and is_xls:
            return self._extract_structured_xls(file_path, doc_name)
        else:
            logger.warning(f"Нет библиотеки для Excel: xlsx={is_xlsx}, xls={is_xls}")
            return ""

    def _detect_format(self, file_path: Path) -> tuple:
        """Определяет формат Excel по магическим байтам."""
        with open(file_path, "rb") as f:
            header = f.read(8)
        is_xlsx = header.startswith(b"\x50\x4b\x03\x04")
        is_xls = header.startswith(b"\xd0\xcf\x11\xe0")

        if not is_xlsx and not is_xls:
            ext = file_path.suffix.lower()
            if ext == ".xlsx":
                is_xlsx = True
            elif ext == ".xls":
                is_xls = True
        return is_xlsx, is_xls

    def _extract_structured_xlsx(self, file_path: Path, doc_name: str) -> str:
        """Структурированное извлечение XLSX."""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_texts = []

        for sheet in wb.worksheets:
            sheet_texts, headers, qty_idx, svc_idx = [], [], -1, -1

            for row_idx, row in enumerate(sheet.iter_rows()):
                row_values = [str(c.value) if c.value is not None else "" for c in row]
                row_text = [v for v in row_values if v.strip()]
                if not row_text:
                    continue

                if row_idx == 0 or (row_idx < 3 and not headers):
                    headers = [v.lower().strip() for v in row_values]
                    qty_idx, svc_idx = self._find_key_columns(headers)
                    continue

                enriched = self._enrich_row(row_values, headers, qty_idx)
                if enriched:
                    sheet_texts.append(" | ".join(enriched))

            # Поиск количества по ключевым словам услуги
            if qty_idx >= 0 and svc_idx >= 0:
                self._inject_quantity_from_sheet(sheet, sheet_texts, qty_idx, svc_idx)

            if sheet_texts:
                all_texts.append(f"=== ЛИСТ: {sheet.title} ===")
                all_texts.extend(sheet_texts)

        result = "\n".join(all_texts)
        logger.info(f"[Excel] Извлечено: {len(all_texts)} строк, {len(result)} символов")
        return result

    def _extract_structured_xls(self, file_path: Path, doc_name: str) -> str:
        """Структурированное извлечение XLS."""
        wb = xlrd.open_workbook(file_path)
        all_texts = []

        for sheet in wb.sheets():
            sheet_texts, headers, qty_idx, svc_idx = [], [], -1, -1

            for row_idx in range(sheet.nrows):
                row_values = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
                row_text = [v for v in row_values if v.strip()]
                if not row_text:
                    continue

                if row_idx == 0 or (row_idx < 3 and not headers):
                    headers = [v.lower().strip() for v in row_values]
                    qty_idx, svc_idx = self._find_key_columns(headers)
                    continue

                enriched = self._enrich_row(row_values, headers, qty_idx)
                if enriched:
                    sheet_texts.append(" | ".join(enriched))

            if qty_idx >= 0 and svc_idx >= 0:
                self._inject_quantity_from_xls_sheet(sheet, sheet_texts, qty_idx, svc_idx)

            if sheet_texts:
                all_texts.append(f"=== ЛИСТ: {sheet.name} ===")
                all_texts.extend(sheet_texts)

        result = "\n".join(all_texts)
        logger.info(f"[Excel XLS] Извлечено: {len(all_texts)} строк, {len(result)} символов")
        return result

    def _find_key_columns(self, headers: list) -> tuple:
        """Находит колонки количества и услуги."""
        qty_idx, svc_idx = -1, -1
        for idx, h in enumerate(headers):
            if any(kw in h for kw in QUANTITY_KEYWORDS):
                qty_idx = idx
                logger.info(f"[Excel] Колонка количества: '{h}' (индекс {idx})")
            if any(kw in h for kw in ["наименование", "услуга", "работа", "предмет"]):
                svc_idx = idx
        return qty_idx, svc_idx

    def _enrich_row(self, row_values: list, headers: list, qty_idx: int) -> list:
        """Добавляет заголовки к ячейкам строки."""
        result = []
        for col_idx, val in enumerate(row_values):
            if not val.strip():
                continue
            if col_idx < len(headers) and headers[col_idx]:
                header = headers[col_idx]
                if col_idx == qty_idx:
                    result.append(f"{header}: {val}")
                else:
                    result.append(val)
            else:
                result.append(val)
        return result

    def _inject_quantity_from_sheet(self, sheet, sheet_texts: list, qty_idx: int, svc_idx: int):
        """Ищет количество по ключевым словам услуги и вставляет в начало."""
        for row in sheet.iter_rows(min_row=2):
            svc_cell = row[svc_idx].value if svc_idx < len(row) else None
            qty_cell = row[qty_idx].value if qty_idx < len(row) else None
            if svc_cell and qty_cell:
                svc_str = str(svc_cell).lower()
                if any(kw in svc_str for kw in SERVICE_KEYWORDS):
                    try:
                        qty = int(float(str(qty_cell).replace(" ", "").replace(",", ".")))
                        sheet_texts.insert(0, f"=== ИЗВЛЕЧЕНО ИЗ ТАБЛИЦЫ: Количество = {qty} ===")
                        logger.info(f"[Excel] Найдено количество: {qty}")
                    except (ValueError, TypeError):
                        pass

    def _inject_quantity_from_xls_sheet(self, sheet, sheet_texts: list, qty_idx: int, svc_idx: int):
        """Аналогично для XLS."""
        for row_idx in range(1, sheet.nrows):
            svc_cell = sheet.cell_value(row_idx, svc_idx) if svc_idx < sheet.ncols else None
            qty_cell = sheet.cell_value(row_idx, qty_idx) if qty_idx < sheet.ncols else None
            if svc_cell and qty_cell:
                svc_str = str(svc_cell).lower()
                if any(kw in svc_str for kw in SERVICE_KEYWORDS):
                    try:
                        qty = int(float(str(qty_cell).replace(" ", "").replace(",", ".")))
                        sheet_texts.insert(0, f"=== ИЗВЛЕЧЕНО ИЗ ТАБЛИЦЫ: Количество = {qty} ===")
                    except (ValueError, TypeError):
                        pass

    def _extract_fallback(self, file_path: Path) -> str:
        """Fallback: обычное извлечение как plain text."""
        try:
            if HAS_OPENPYXL:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                texts = []
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows():
                        row_text = [str(c.value) for c in row if c.value is not None]
                        if row_text:
                            texts.append(" | ".join(row_text))
                return "\n".join(texts)
            elif HAS_XLRD:
                wb = xlrd.open_workbook(file_path)
                texts = []
                for sheet in wb.sheets():
                    for row_idx in range(sheet.nrows):
                        row_text = [str(sheet.cell_value(row_idx, c)) for c in range(sheet.ncols)]
                        if row_text:
                            texts.append(" | ".join(row_text))
                return "\n".join(texts)
            return ""
        except Exception as e:
            logger.error(f"Ошибка fallback Excel: {e}")
            return ""
